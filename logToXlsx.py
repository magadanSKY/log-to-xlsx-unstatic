import openpyxl
import re


with open('Data\\access.2023-08-19.log', 'r') as file: #выбор лог файла для парса
    content = file.read()

lines = content.split('\n')

#задаём ключи для поиска
counter = {}
for line in lines[:-1]:
    cells = line.split(' ')

    date = cells[3][1:cells[3].find(':')]

    time = cells[3][13:21]
    time_index = time.find(':')
    time = int(time[:time_index])

#прописываем форматирование
    url = "root" if not cells[6][0:] else cells[6][0:].partition('?')[0]
    url = re.sub(r'product/\d+', 'product/#sku', url)
    url = re.sub(r'stores/\w\d+', 'stores/#store', url)
    url = re.sub(r'catalog/\D+', 'catalog/nameOfCatalog/', url)
    url = re.sub(r'brands/\D+', 'brands/nameOfBrand', url)
    url = re.sub(r'blog/\D+', 'blog/nameOfBlog', url)

    if (".js" in url) or (".css" in url) or (".ico" in url) or (".png" in url) or (".webp" in url) or (".svg" in url) or ("Wx" in url) or (".jpg" in url) or (".woff2" in url) or (".ttf" in url) or (".woff" in url) or (".gif" in url):
        pass
    else:
        if date in counter:
            if url in counter[date]:
                if time in counter[date][url]:
                    counter[date][url][time] += 1
                else:
                    counter[date][url][time] = 1
            else:
                counter[date][url] = {time:  1}
        else:
            counter[date] = {url: {time:  1}}
workbook = openpyxl.Workbook()

#выводим все в .xlsx
sheet = workbook.active

sheet['A1'] = "Operation"
sheet['B1'] = "Operation_count"
sheet['C1'] = "Date"
sheet['D1'] = "Hour"

row_num = 2
for date, value in counter.items():
    for url, value1 in value.items():
        for time, count in value1.items():
            sheet.cell(row=row_num, column=1).value = url
            sheet.cell(row=row_num, column=2).value = count
            sheet.cell(row=row_num, column=3).value = date
            sheet.cell(row=row_num, column=4).value = time
            row_num += 1

workbook.save('Result\\outputTest9.xlsx')